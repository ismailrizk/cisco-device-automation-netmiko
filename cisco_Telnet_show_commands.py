import openpyxl
from netmiko import ConnectHandler
import logging
from openpyxl.styles import PatternFill
import os

logging.basicConfig(
    filename='telnet_logs.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

OUTPUT_DIR = "device_outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

COMMANDS = [
    'show ip interface brief',
    'show interface status',
    'show cdp neighbors',
    'show cdp neighbors detail',
    'show vrf',
    'show vlan',
    'show ip route',
    'show ip ospf neighbor',
    'show ip ospf interface brief',
    'show ip route 192.168.40.1',
    'show running-config'
]


def export_device_info(ip, username, password, output_file):

    device = {
        'device_type': 'cisco_ios_telnet',
        'ip': ip,
        'username': username,
        'password': password,
        'port': 23,
    #    'timeout': 10,
    #    'delay_factor': 2,
    }

    try:
        with ConnectHandler(**device) as net_connect:
            logging.info(f"Connected to {ip} using {username}")

            outputs = {}

            for command in COMMANDS:
                outputs[command] = net_connect.send_command(command)

            with open(output_file, 'w') as file:
                for command, output in outputs.items():
                    file.write(f"\n{'='*20}\n")
                    file.write(f"{command}\n")
                    file.write(f"{'='*20}\n")
                    file.write(output + "\n")

        return True

    except Exception as e:
        logging.error(f"Failed to connect to {ip} with {username}: {e}")
        return False


if __name__ == "__main__":

    excel_path = input("Enter the path to the Excel sheet: ").strip()
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    credentials = {
        'cisco': 'cisco',
        'admin': 'Admin@123'
    }

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):

        ip = row[0]
        if not ip:
            continue

        output_file = os.path.join(OUTPUT_DIR, f"{ip}.txt")

        for username, password in credentials.items():
            print(f"Trying {ip} with {username}")
            if export_device_info(ip, username, password, output_file):
                break
        else:
            cell = sheet.cell(row=row_index, column=1)
            cell.fill = PatternFill(
                start_color='FFFF0000',
                end_color='FFFF0000',
                fill_type='solid'
            )

    wb.save(excel_path)

